import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import os
from openpyxl import load_workbook

# 設定 PTT 看板的 URL (這裡以 Gossiping 板為例)
base_url = "https://www.ptt.cc"
board_url = "/bbs/Gossiping/index.html"
cookies = {'over18': '1'}  # PTT 需要加入年齡確認的 cookie

# 指定保存 Excel 檔案的路徑
save_path = "C:\\Users\\T14 Gen 3\\OneDrive\\個人實驗室\\my_ptt_posts.xlsx"

# 儲存所有文章的標題、連結和時間
data = []

# 設定爬取的頁數 (可以自行調整)
max_pages = 5
current_page = 0

def get_post_content(link):
    # 發送 GET 請求並加入 cookies
    response = requests.get(link, cookies=cookies)
    
    # 確保請求成功
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        # 取得文章內容
        content = soup.find('div', id='main-content')
        if content:
            # 移除不需要的標籤（文章下方的推文等）
            for tag in content.find_all(['span', 'div']):
                tag.extract()
            return content.text.strip()
    return None

# 如果已經有檔案存在，讀取已爬取的資料
if os.path.exists(save_path):
    df_existing = pd.read_excel(save_path)
    existing_titles = set(df_existing['標題'])  # 用 set 加速比對
else:
    existing_titles = set()  # 如果檔案不存在，表示沒有已經爬取的資料

while current_page < max_pages:
    # 發送 GET 請求
    response = requests.get(base_url + board_url, cookies=cookies)
    if response.status_code == 200:
        # 使用 BeautifulSoup 解析 HTML
        soup = BeautifulSoup(response.text, 'html.parser')

        # 找到所有文章標題的標籤
        posts = soup.find_all('div', class_='r-ent')

        for post in posts:
            a_tag = post.find('a')
            if a_tag:
                title = a_tag.text.strip()  # 標題
                link = base_url + a_tag['href']  # 連結
                # 測試抓取某篇文章的內容
                content = get_post_content(link)

                # 擷取發佈時間
                date = post.find('div', class_='date').text.strip()  # 發佈日期
                if title not in existing_titles:
                    data.append([date, title, content, link])
                    existing_titles.add(title)  # 加入已保存的標題集合

        # 找到「上一頁」的連結
        btn_group = soup.find('div', class_='btn-group btn-group-paging')
        prev_link = btn_group.find_all('a')[1]['href']  # 第二個 a 標籤為「上一頁」
        board_url = prev_link

        # 增加當前頁數
        current_page += 1

        # 避免過快爬取，設定延遲
        time.sleep(2)
    else:
        print(f"無法獲取頁面，狀態碼: {response.status_code}")
        break

# 將新爬取的資料轉換為 pandas DataFrame
if data:
    df_new = pd.DataFrame(data, columns=["發佈日期", "標題", "內容", "連結"])

    # 如果檔案已經存在，則附加新資料
    if os.path.exists(save_path):
        # 使用 openpyxl 來追加資料
        with pd.ExcelWriter(save_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # 獲取當前表單的最大行數，追加資料時從該行開始
            existing_workbook = load_workbook(save_path)
            sheet = existing_workbook.active
            max_row = sheet.max_row

            df_new.to_excel(writer, index=False, header=False, startrow=max_row)
    else:
        # 如果檔案不存在，則直接創建新檔案並寫入資料
        df_new.to_excel(save_path, index=False, engine='openpyxl')

    print(f"成功將新資料追加至 {save_path}")
else:
    print("沒有新資料可以追加。")